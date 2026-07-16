#!/usr/bin/env python3
"""
Build Rossby Financial's Multi-Office CVP (Cost-Volume-Profit) Excel model.

This generates `Rossby_CVP_Multi_Office_Model.xlsx` at the repo root. Every computed
cell is a live Excel formula (not a baked value) so the finance team can edit
assumptions in-sheet and watch breakeven move.

Data sources:
  - P&L: Rossby Financial LLC Profit and Loss, Jan 1 - Jul 1 2026 (6 months, cash basis)
  - Office book: Rossby 2026 Billing workbook, tab "2026 Billing" (18 active offices)
  - Pricing engine: ported from index.html (rossbyPricing / rossbyPerAccountRate /
    additionalAdvisorFee), so this internal model and the advisor-facing calculator agree.

Run:  python3 tools/build_cvp_model.py
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import LineChart, Reference, Series
from openpyxl.utils import get_column_letter

# ----------------------------------------------------------------------------
# Brand palette (Rossby Maritime v2)
# ----------------------------------------------------------------------------
NAVY = "23286C"
NAVY_DEEP = "1A1E52"
GOLD = "CBA96A"
GOLD_DEEP = "B8924E"
CREAM = "F7F3EC"
CANVAS = "FDFCF8"
INK = "1B1B25"
GREY = "6B6B78"
OK_GREEN = "1E7A4D"
WARN_RED = "B23A2E"
PANEL = "EFEBE1"

WHITE = "FFFFFF"

def fill(hexcolor):
    return PatternFill("solid", fgColor=hexcolor)

thin = Side(style="thin", color="D9D2C4")
med = Side(style="medium", color=GOLD_DEEP)
box = Border(left=thin, right=thin, top=thin, bottom=thin)
gold_top = Border(top=Side(style="medium", color=GOLD))

def style_cell(c, *, bold=False, size=11, color=INK, bg=None, italic=False,
               align=None, valign="center", wrap=False, number_format=None,
               border=None, name="Calibri"):
    c.font = Font(name=name, bold=bold, size=size, color=color, italic=italic)
    if bg:
        c.fill = fill(bg)
    c.alignment = Alignment(horizontal=align, vertical=valign, wrap_text=wrap)
    if number_format:
        c.number_format = number_format
    if border:
        c.border = border
    return c

USD = '_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)'
USD2 = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"_);_(@_)'
NUM0 = '#,##0'
NUM2 = '#,##0.00'
PCT = '0.0%'
PCT2 = '0.00%'

# ============================================================================
# Source data
# ============================================================================

# --- P&L fixed-cost lines (6-month actuals). include/one-time flags per plan ---
# (label, 6mo_actual, include_default, note)
FIXED_LINES = [
    ("Advertising & Marketing", 15281.00, "Y", "Recurring"),
    ("Bank Fees", 3157.35, "Y", "Recurring"),
    ("Contractor Expense - G&A", 1272.50, "Y", "Recurring"),
    ("Gifts", 205.98, "Y", "Recurring"),
    ("Insurance Expense", 10094.85, "Y", "Recurring"),
    ("Interest Expense", 22936.81, "N", "REMOVE - debt paid off (gone from books)"),
    ("Licensing Fees", 2686.00, "Y", "Recurring"),
    ("Meals & Entertainment", 1392.34, "Y", "Recurring"),
    ("Office Expenses", 616.47, "Y", "Recurring"),
    ("Payroll Expenses (401k, taxes, wages)", 91284.53, "Y", "Recurring - scales in steps"),
    ("Professional Services (Legal + Tax)", 4410.00, "Y", "Recurring"),
    ("Public Relations Expense", 1900.00, "Y", "Recurring"),
    ("QuickBooks Payments Fees", 5656.61, "Y", "Recurring"),
    ("Regulatory Fees", 1950.00, "Y", "Recurring"),
    ("Rent Expense", 3147.00, "Y", "Recurring"),
    ("Software (Advisor Platform, Back Office, Compliance)", 119801.14, "Y", "Recurring - step-fixed w/ scale"),
    ("Suspense", 500.00, "N", "FLAG - unclassified / one-time"),
    ("Travel Expenses", 19799.88, "Y", "Recurring - lumpy"),
    ("Conference Expense", 11783.65, "N", "FLAG - likely one-time / lumpy"),
    ("Utilities Expense", 4312.95, "Y", "Recurring"),
]

# --- Pricing engine tiers (ported from index.html) ---
# per-account $/month by minimum account count (descending threshold, decreasing rate)
PER_ACCT_TIERS = [
    (10001, 0.75),
    (5501, 1.00),
    (3401, 2.00),
    (2101, 3.00),
    (1301, 4.00),
    (501, 5.00),
    (101, 6.00),
    (0, 7.00),
]
# per additional-advisor $/year by minimum advisor headcount
ADVISOR_TIERS = [
    (14, 1000),
    (9, 1500),
    (6, 2500),
    (4, 5000),
    (0, 7500),
]

# --- 19 real offices from the 2026 Billing tab ---
# (name, model, base_fee, accounts, iars, per_acct_billed, per_advisor_billed, discount, actual_total)
# accounts=None where the billing sheet had "N/A" (Founders flat-fee offices)
OFFICES = [
    ("Red Oak",               "Founders", 25000, None,  1,     0,     0, 0.00, 25000),
    ("Life Strategies",       "Founders", 25000, None,  2,     0,     0, 0.00, 25000),
    ("EViE",                  "Founders", 25000, None,  2,     0,  5000, 0.00, 30000),
    ("Florio",                "Founders", 25000, None,  1,     0,     0, 0.00, 25000),
    ("Toler",                 "Access",   20000,  854,  5, 40992, 20000, 0.00, 80992),
    ("RMRG",                  "Access",   20000,  124,  2,  7440,  7500, 0.20, 27952),
    ("Trucess",               "Access",   15000,  130,  1,  7800,     0, 0.00, 22800),
    ("Reymar Wealth",         "Access",   20000,  131,  1,  7860,     0, 0.00, 27860),
    ("Alan R. Joyce Co.",     "Access",   20000,   14,  1,  1008,     0, 0.00, 21008),
    ("OSBORN Wealth",         "Access",   20000,  197,  2,  9600,  7500, 0.00, 37100),
    ("Windward Financial",    "Access",   20000,  119,  1,  7140,     0, 0.00, 27140),
    ("Financial Strategies",  "Access",   20000,  319,  1, 19140,     0, 0.00, 39140),
    ("Prarie Oak",            "Access",   20000,  231,  2, 13860,  7500, 0.00, 41360),
    ("Sequence Wealth",       "Access",   20000,   59,  1,  4248,     0, 0.00, 24248),
    ("Oak Stream Investments","Access",   20000,   88,  1,  7392,     0, 0.00, 27392),
    ("Tribox",                "Access",   20000,  261,  2, 18792,  7500, 0.00, 46292),
    ("John Lyon",             "Access",   20000,  300,  1, 21600,     0, 0.00, 41600),
    ("AP Financial",          "Access",   20000,   20,  1,  1680,     0, 0.00, 21680),
]

BLANK_PROSPECT_ROWS = 11  # empty formula'd rows for new offices


# ============================================================================
# Workbook
# ============================================================================
wb = Workbook()

# --------------------------------------------------------------------------
# Helper for section header rows
# --------------------------------------------------------------------------
def section_banner(ws, row, first_col, last_col, text, *, bg=NAVY, fg=WHITE, size=12):
    ws.merge_cells(start_row=row, start_column=first_col, end_row=row, end_column=last_col)
    c = ws.cell(row, first_col, text)
    style_cell(c, bold=True, size=size, color=fg, bg=bg, align="left")
    ws.row_dimensions[row].height = 24
    for col in range(first_col, last_col + 1):
        ws.cell(row, col).fill = fill(bg)


# ============================================================================
# TAB 1 — Read Me
# ============================================================================
ws1 = wb.active
ws1.title = "Read Me"
ws1.sheet_view.showGridLines = False
ws1.column_dimensions["A"].width = 3
ws1.column_dimensions["B"].width = 104

def rm(row, text, *, bold=False, size=11, color=INK, bg=None, indent=0, size_row=None):
    c = ws1.cell(row, 2, text)
    style_cell(c, bold=bold, size=size, color=color, bg=bg, align="left", wrap=True, valign="top")
    if indent:
        c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=indent)
    if size_row:
        ws1.row_dimensions[row].height = size_row
    return c

ws1.merge_cells("B1:B1")
style_cell(ws1.cell(1, 2, "ROSSBY FINANCIAL"), bold=True, size=20, color=NAVY, align="left")
ws1.row_dimensions[1].height = 28
style_cell(ws1.cell(2, 2, "Multi-Office CVP (Cost-Volume-Profit) & Breakeven Model"),
           bold=True, size=14, color=GOLD_DEEP, align="left")
ws1.row_dimensions[2].height = 22

rows_content = [
    (4, "What this model answers", True, 12, NAVY, None),
    (5, "How many advisor offices does Rossby need to cover its firm-wide fixed costs, and how does that "
        "breakeven move as we add offices? It also tracks every existing office's contribution to the firm.",
        False, 11, INK, None),
    (7, "How the tabs fit together", True, 12, NAVY, None),
    (8, "1.  Assumptions & Fixed Costs  —  set the annualization factor and toggle which P&L lines are ongoing "
        "fixed costs. Produces Total Go-Forward Annual Fixed Costs. Holds the editable pricing-engine tiers.",
        False, 11, INK, None),
    (9, "2.  Office Register  —  every office (18 existing, preloaded from the 2026 Billing book) with its "
        "revenue to Rossby and contribution. Add prospects in the blank rows.", False, 11, INK, None),
    (10, "3.  CVP Dashboard  —  the headline: total contribution vs fixed costs, firm net income, breakeven "
         "office count, and how the per-office breakeven bar falls as offices are added.", False, 11, INK, None),
    (11, "4.  New Office Evaluator  —  price a prospective office off the tiered engine and see its marginal "
         "effect on firm net income and breakeven.", False, 11, INK, None),
    (13, "Key definitions", True, 12, NAVY, None),
    (14, "Contribution (per office)  =  Revenue to Rossby  −  Cost to Serve.  This is what each office contributes "
         "toward covering firm fixed costs. Rossby's revenue is platform/billing revenue (base + per-account + "
         "per-advisor fees), NOT the advisory fees that pass through to advisors.", False, 11, INK, None),
    (16, "Fixed Costs  =  Rossby's own overhead (software, payroll, rent, compliance, etc.) that does not scale "
         "one-for-one with each office. Taken from the P&L, annualized, with the paid-off debt and one-time items "
         "removed via the toggles on Tab 2.", False, 11, INK, None),
    (18, "Breakeven # of offices  =  Fixed Costs  ÷  Average Contribution per Office.  The per-office breakeven "
         "THRESHOLD (Fixed ÷ N) falls as you add offices, because the fixed base is shared across more offices.",
         False, 11, INK, None),
    (20, "Data sources", True, 12, NAVY, None),
    (21, "•  P&L: Rossby Financial LLC Profit & Loss, Jan 1 – Jul 1 2026 (6 months, cash basis).", False, 11, GREY, None),
    (22, "•  Offices: Rossby 2026 Billing workbook, tab \"2026 Billing\" (18 active offices, 2,847 accounts, 28 IARs).",
         False, 11, GREY, None),
    (23, "•  Pricing engine (base + per-account + per-advisor tiers) is ported from the advisor-facing calculator "
         "(index.html) so both tools price offices identically.", False, 11, GREY, None),
    (25, "Editing & assumptions", True, 12, NAVY, None),
    (26, "Blue-tinted cells are inputs — edit them freely. Everything else is a live formula and recalculates. "
         "The annualization factor (Tab 2) defaults to 2 (×2 to turn the 6-month P&L into an annual run-rate); "
         "change it to 1 to work in the actual 6-month period.", False, 11, INK, None),
    (28, "Reconciliation check", True, 12, NAVY, None),
    (29, "To tie back to the P&L: on Tab 2 set the factor to 1 and switch Interest, Suspense and Conference to \"Y\". "
         "Total fixed then equals the P&L's $322,189 operating expense.", False, 11, INK, None),
]
for r, text, bold, size, color, bg in rows_content:
    rm(r, text, bold=bold, size=size, color=color, bg=bg)

# input-legend swatch
style_cell(ws1.cell(31, 2, "  Input cell (edit me)"), bg="DCE6F5", align="left", color=NAVY, bold=True)

for r in [5, 8, 9, 10, 11, 14, 16, 18, 21, 22, 23, 26, 29]:
    ws1.row_dimensions[r].height = 30

# ============================================================================
# TAB 2 — Assumptions & Fixed Costs
# ============================================================================
ws2 = wb.create_sheet("Assumptions & Fixed Costs")
ws2.sheet_view.showGridLines = False
widths2 = {"A": 2.5, "B": 46, "C": 15, "D": 11, "E": 14, "F": 16, "G": 42}
for col, w in widths2.items():
    ws2.column_dimensions[col].width = w

style_cell(ws2.cell(1, 2, "Assumptions & Fixed Costs"), bold=True, size=16, color=NAVY, align="left")
ws2.row_dimensions[1].height = 24

# --- Annualization control ---
style_cell(ws2.cell(3, 2, "P&L period"), bold=True, align="left")
style_cell(ws2.cell(3, 3, "Jan 1 – Jul 1 2026 (6 mo)"), align="left", color=GREY)
style_cell(ws2.cell(4, 2, "Annualization factor"), bold=True, align="left")
fac = ws2.cell(4, 3, 2)
style_cell(fac, bold=True, size=12, color=NAVY, bg="DCE6F5", align="center", border=box)
FACTOR_REF = "'Assumptions & Fixed Costs'!$C$4"
style_cell(ws2.cell(4, 4, "×  (2 = annualize; 1 = keep 6-mo actuals)"), align="left", color=GREY, italic=True)

# --- Fixed cost table ---
hdr = 6
section_banner(ws2, hdr, 2, 7, "GO-FORWARD FIXED COSTS  (Rossby firm overhead)")
hrow = hdr + 1
headers2 = ["Expense line", "6-mo Actual", "Include?", "Recurring?", "Annualized", "Note"]
for i, h in enumerate(headers2):
    style_cell(ws2.cell(hrow, 2 + i, h), bold=True, color=WHITE, bg=NAVY_DEEP,
               align="center" if i else "left", border=box)

dv_yn = DataValidation(type="list", formula1='"Y,N"', allow_blank=False)
ws2.add_data_validation(dv_yn)

first_data = hrow + 1
for j, (label, amt, inc, note) in enumerate(FIXED_LINES):
    r = first_data + j
    style_cell(ws2.cell(r, 2, label), align="left", border=box)
    style_cell(ws2.cell(r, 3, amt), number_format=USD2, align="right", border=box)
    inc_cell = ws2.cell(r, 4, inc)
    style_cell(inc_cell, align="center", bg="DCE6F5", bold=True,
               color=OK_GREEN if inc == "Y" else WARN_RED, border=box)
    dv_yn.add(inc_cell)
    style_cell(ws2.cell(r, 5, "Y"), align="center", color=GREY, border=box)
    # Annualized = actual * factor * (include?1:0)
    ann = ws2.cell(r, 6)
    ann.value = f'=C{r}*{FACTOR_REF}*IF(D{r}="Y",1,0)'
    style_cell(ann, number_format=USD, align="right", border=box)
    note_color = WARN_RED if note.startswith(("REMOVE", "FLAG")) else GREY
    style_cell(ws2.cell(r, 7, note), align="left", color=note_color,
               italic=True, border=box, size=10)

last_data = first_data + len(FIXED_LINES) - 1
tot_row = last_data + 1
style_cell(ws2.cell(tot_row, 2, "TOTAL GO-FORWARD ANNUAL FIXED COSTS"), bold=True,
           color=NAVY, bg=CREAM, align="left", border=gold_top)
style_cell(ws2.cell(tot_row, 3), bg=CREAM, border=gold_top)
style_cell(ws2.cell(tot_row, 4), bg=CREAM, border=gold_top)
style_cell(ws2.cell(tot_row, 5), bg=CREAM, border=gold_top)
tf = ws2.cell(tot_row, 6, f"=SUM(F{first_data}:F{last_data})")
style_cell(tf, bold=True, size=12, color=NAVY, bg=CREAM, number_format=USD, align="right", border=gold_top)
style_cell(ws2.cell(tot_row, 7), bg=CREAM, border=gold_top)
FIXED_TOTAL_REF = f"'Assumptions & Fixed Costs'!$F${tot_row}"

# 6-mo included subtotal helper line
sub_row = tot_row + 1
style_cell(ws2.cell(sub_row, 2, "(memo) 6-mo included subtotal"), italic=True, color=GREY, align="left")
style_cell(ws2.cell(sub_row, 6,
           f'=SUMPRODUCT((D{first_data}:D{last_data}="Y")*C{first_data}:C{last_data})'),
           number_format=USD, align="right", color=GREY, italic=True)

# --- Pricing engine tiers ---
eng = sub_row + 3
section_banner(ws2, eng, 2, 7, "PRICING ENGINE  (revenue to Rossby per office — editable list-price grid)")

# base fees
br = eng + 1
style_cell(ws2.cell(br, 2, "Base platform fee — Access model"), align="left", bold=True)
bf_access = ws2.cell(br, 3, 20000)
style_cell(bf_access, number_format=USD, align="right", bg="DCE6F5", border=box)
style_cell(ws2.cell(br + 1, 2, "Base platform fee — Founders model"), align="left", bold=True)
bf_found = ws2.cell(br + 1, 3, 25000)
style_cell(bf_found, number_format=USD, align="right", bg="DCE6F5", border=box)
BF_ACCESS_REF = f"'Assumptions & Fixed Costs'!$C${br}"
BF_FOUND_REF = f"'Assumptions & Fixed Costs'!$C${br+1}"

# per-account tier table
pa_hdr = br + 3
style_cell(ws2.cell(pa_hdr, 2, "Per-account rate ($/account/month)"), bold=True, color=NAVY, align="left")
style_cell(ws2.cell(pa_hdr + 1, 2, "Min accounts (>=)"), bold=True, bg=NAVY_DEEP, color=WHITE, align="center", border=box)
style_cell(ws2.cell(pa_hdr + 1, 3, "$/acct/mo"), bold=True, bg=NAVY_DEEP, color=WHITE, align="center", border=box)
# store ascending by threshold for LOOKUP (LOOKUP needs ascending lookup vector)
pa_tiers_asc = sorted(PER_ACCT_TIERS, key=lambda t: t[0])
pa_first = pa_hdr + 2
for k, (thr, rate) in enumerate(pa_tiers_asc):
    r = pa_first + k
    style_cell(ws2.cell(r, 2, thr), number_format=NUM0, align="center", bg="DCE6F5", border=box)
    style_cell(ws2.cell(r, 3, rate), number_format=NUM2, align="center", bg="DCE6F5", border=box)
pa_last = pa_first + len(pa_tiers_asc) - 1
PA_LOOKUP_VEC = f"'Assumptions & Fixed Costs'!$B${pa_first}:$B${pa_last}"
PA_RESULT_VEC = f"'Assumptions & Fixed Costs'!$C${pa_first}:$C${pa_last}"

# per-advisor tier table
ad_hdr = pa_last + 2
style_cell(ws2.cell(ad_hdr, 2, "Per additional-advisor fee ($/extra advisor/year)"), bold=True, color=NAVY, align="left")
style_cell(ws2.cell(ad_hdr + 1, 2, "Min advisors (>=)"), bold=True, bg=NAVY_DEEP, color=WHITE, align="center", border=box)
style_cell(ws2.cell(ad_hdr + 1, 3, "$/extra advisor/yr"), bold=True, bg=NAVY_DEEP, color=WHITE, align="center", border=box)
ad_tiers_asc = sorted(ADVISOR_TIERS, key=lambda t: t[0])
ad_first = ad_hdr + 2
for k, (thr, rate) in enumerate(ad_tiers_asc):
    r = ad_first + k
    style_cell(ws2.cell(r, 2, thr), number_format=NUM0, align="center", bg="DCE6F5", border=box)
    style_cell(ws2.cell(r, 3, rate), number_format=USD, align="center", bg="DCE6F5", border=box)
ad_last = ad_first + len(ad_tiers_asc) - 1
AD_LOOKUP_VEC = f"'Assumptions & Fixed Costs'!$B${ad_first}:$B${ad_last}"
AD_RESULT_VEC = f"'Assumptions & Fixed Costs'!$C${ad_first}:$C${ad_last}"

# optional override channel
ov = ad_last + 2
style_cell(ws2.cell(ov, 2, "Optional: advisory override keep %  (2nd revenue channel, off by default)"),
           align="left", bold=True)
ov_cell = ws2.cell(ov, 3, 0.0)
style_cell(ov_cell, number_format=PCT2, align="right", bg="DCE6F5", border=box)
style_cell(ws2.cell(ov + 1, 2, "   Applied as AUM × bps/10000 × keep%. Leave 0 to model platform fees only."),
           align="left", color=GREY, italic=True, size=10)

# ============================================================================
# TAB 3 — Office Register
# ============================================================================
ws3 = wb.create_sheet("Office Register")
ws3.sheet_view.showGridLines = False
cols3 = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N"]
widths3 = [2.5, 22, 11, 11, 10, 8, 11, 12, 10, 16, 16, 11, 15, 15]
for col, w in zip(cols3, widths3):
    ws3.column_dimensions[col].width = w

style_cell(ws3.cell(1, 2, "Office Register"), bold=True, size=16, color=NAVY, align="left")
ws3.row_dimensions[1].height = 24
style_cell(ws3.cell(2, 2, "Existing offices preloaded from the 2026 Billing book. Add prospects in the blank rows below."),
           align="left", color=GREY, italic=True)

h3 = 4
headers3 = [
    ("Office", "left"),
    ("Status", "center"),
    ("Model", "center"),
    ("Base Fee", "right"),
    ("Accounts", "center"),
    ("IARs", "center"),
    ("Per-Acct $/mo", "right"),
    ("Per-Advisor $", "right"),
    ("Discount %", "center"),
    ("Actual Annual Rev", "right"),
    ("Modeled (List)", "right"),
    ("Rev Used", "center"),
    ("Cost to Serve", "right"),
    ("Contribution", "right"),
]
for i, (h, al) in enumerate(headers3):
    style_cell(ws3.cell(h3, 2 + i, h), bold=True, color=WHITE, bg=NAVY_DEEP, align=al,
               wrap=True, border=box, size=10)
ws3.row_dimensions[h3].height = 30

dv_status = DataValidation(type="list", formula1='"Existing,Prospect"', allow_blank=True)
dv_model = DataValidation(type="list", formula1='"Access,Founders"', allow_blank=True)
dv_revused = DataValidation(type="list", formula1='"Actual,Modeled"', allow_blank=True)
for dv in (dv_status, dv_model, dv_revused):
    ws3.add_data_validation(dv)

data_start3 = h3 + 1

def office_row(ws, r, office=None):
    """Write one office row. office=None => blank prospect row (formulas still live)."""
    input_bg = "DCE6F5"
    # Columns:
    # B name | C status | D model | E base | F accounts | G iars | H per-acct | I per-adv | J disc
    # K actual | L modeled | M revused | N cost-serve | O contribution
    if office:
        name, model, base, accts, iars, pa_bill, adv_bill, disc, actual = office
        status = "Existing"
        revused = "Actual"
    else:
        name, model, base, accts, iars, pa_bill, adv_bill, disc, actual = ("", "", "", "", "", "", "", "", "")
        status, revused = "", "Modeled"

    style_cell(ws.cell(r, 2, name), align="left", bg=input_bg, border=box)
    sc = ws.cell(r, 3, status); style_cell(sc, align="center", bg=input_bg, border=box); dv_status.add(sc)
    mc = ws.cell(r, 4, model); style_cell(mc, align="center", bg=input_bg, border=box); dv_model.add(mc)
    style_cell(ws.cell(r, 5, base), number_format=USD, align="right", bg=input_bg, border=box)
    style_cell(ws.cell(r, 6, accts), number_format=NUM0, align="center", bg=input_bg, border=box)
    style_cell(ws.cell(r, 7, iars), number_format=NUM0, align="center", bg=input_bg, border=box)
    # actual per-account $/mo billed (for existing) and per-advisor $ billed
    style_cell(ws.cell(r, 8, (pa_bill / accts / 12) if (office and accts) else ""),
               number_format=NUM2, align="right", bg=input_bg, border=box)
    style_cell(ws.cell(r, 9, adv_bill if office else ""), number_format=USD, align="right", bg=input_bg, border=box)
    style_cell(ws.cell(r, 10, disc if office else ""), number_format=PCT, align="center", bg=input_bg, border=box)
    # Actual annual revenue to Rossby (input for existing)
    style_cell(ws.cell(r, 11, actual if office else ""), number_format=USD, align="right", bg=input_bg, border=box)

    # Modeled list-price revenue (live formula off the engine tiers)
    # base_model = IF(model="Founders", BF_FOUND, BF_ACCESS) when base blank; else use entered base
    # per-acct component = accounts * LOOKUP(accounts, PA vec) * 12  (Access only; Founders have no per-acct)
    # per-advisor = MAX(iars-1,0) * LOOKUP(iars, AD vec)
    # override (optional) currently AUM not tracked per-office here -> 0 unless added
    modeled = (
        f'=IF(B{r}="","",'  # blank prospect rows stay empty
        f'IF(N(F{r})=0,'  # if no accounts -> founders-style flat + advisors
        f'IF(E{r}="",IF(D{r}="Founders",{BF_FOUND_REF},{BF_ACCESS_REF}),E{r})'
        f'+MAX(N(G{r})-1,0)*LOOKUP(N(G{r}),{AD_LOOKUP_VEC},{AD_RESULT_VEC}),'
        f'IF(E{r}="",IF(D{r}="Founders",{BF_FOUND_REF},{BF_ACCESS_REF}),E{r})'
        f'+F{r}*LOOKUP(F{r},{PA_LOOKUP_VEC},{PA_RESULT_VEC})*12'
        f'+MAX(N(G{r})-1,0)*LOOKUP(N(G{r}),{AD_LOOKUP_VEC},{AD_RESULT_VEC})))'
    )
    style_cell(ws.cell(r, 12, modeled), number_format=USD, align="right", border=box, color=GREY)

    rv = ws.cell(r, 13, revused); style_cell(rv, align="center", bg=input_bg, border=box); dv_revused.add(rv)
    style_cell(ws.cell(r, 14, 0 if office else ""), number_format=USD, align="right", bg=input_bg, border=box)

    # Revenue used = actual or modeled; Contribution = rev used - cost to serve
    revused_formula = f'=IF(M{r}="Modeled",L{r},IF(N(K{r})>0,K{r},L{r}))'
    # put "rev used" numeric into a helper column? We'll fold into contribution directly.
    contrib = f'=IF(B{r}="","",{revused_formula[1:]}-N(N{r}))'
    style_cell(ws.cell(r, 15, contrib), number_format=USD, align="right", border=box, bold=True, color=NAVY)

# Widen O col for contribution
ws3.column_dimensions["O"].width = 15

for idx, office in enumerate(OFFICES):
    office_row(ws3, data_start3 + idx, office)
for k in range(BLANK_PROSPECT_ROWS):
    office_row(ws3, data_start3 + len(OFFICES) + k, None)

data_end3 = data_start3 + len(OFFICES) + BLANK_PROSPECT_ROWS - 1

# Totals row
trow3 = data_end3 + 1
style_cell(ws3.cell(trow3, 2, "TOTALS"), bold=True, color=NAVY, bg=CREAM, align="left", border=gold_top)
for col in range(3, 16):
    style_cell(ws3.cell(trow3, col), bg=CREAM, border=gold_top)
# office count (non-blank names)
style_cell(ws3.cell(trow3, 3, f'=COUNTA(B{data_start3}:B{data_end3})'), bold=True, align="center",
           bg=CREAM, border=gold_top, color=NAVY)
style_cell(ws3.cell(trow3, 6, f'=SUM(F{data_start3}:F{data_end3})'), number_format=NUM0, align="center",
           bg=CREAM, border=gold_top)
style_cell(ws3.cell(trow3, 7, f'=SUM(G{data_start3}:G{data_end3})'), number_format=NUM0, align="center",
           bg=CREAM, border=gold_top)
style_cell(ws3.cell(trow3, 14, f'=SUM(N{data_start3}:N{data_end3})'), number_format=USD, align="right",
           bg=CREAM, border=gold_top)
style_cell(ws3.cell(trow3, 15, f'=SUM(O{data_start3}:O{data_end3})'), bold=True, size=12, color=NAVY,
           number_format=USD, align="right", bg=CREAM, border=gold_top)

OFFICE_COUNT_REF = f"'Office Register'!$C${trow3}"
TOTAL_CONTRIB_REF = f"'Office Register'!$O${trow3}"
CONTRIB_RANGE = f"'Office Register'!$O${data_start3}:$O${data_end3}"

# ============================================================================
# TAB 4 — CVP Dashboard & Breakeven
# ============================================================================
ws4 = wb.create_sheet("CVP Dashboard")
ws4.sheet_view.showGridLines = False
widths4 = {"A": 2.5, "B": 40, "C": 18, "D": 4, "E": 14, "F": 16, "G": 16, "H": 16}
for col, w in widths4.items():
    ws4.column_dimensions[col].width = w

style_cell(ws4.cell(1, 2, "CVP Dashboard & Breakeven"), bold=True, size=16, color=NAVY, align="left")
ws4.row_dimensions[1].height = 24

# KPI block
section_banner(ws4, 3, 2, 3, "FIRM POSITION TODAY")
kpis = [
    (4, "Active offices", f"={OFFICE_COUNT_REF}", NUM0),
    (5, "Total annual contribution", f"={TOTAL_CONTRIB_REF}", USD),
    (6, "Total go-forward annual fixed costs", f"={FIXED_TOTAL_REF}", USD),
    (7, "Firm Net Operating Income", f"={TOTAL_CONTRIB_REF}-{FIXED_TOTAL_REF}", USD),
    (8, "Average contribution per office", f"=IF({OFFICE_COUNT_REF}=0,0,{TOTAL_CONTRIB_REF}/{OFFICE_COUNT_REF})", USD),
]
for r, label, formula, fmt in kpis:
    style_cell(ws4.cell(r, 2, label), align="left", border=box)
    c = ws4.cell(r, 3, formula)
    is_ni = (r == 7)
    style_cell(c, number_format=fmt, align="right", border=box, bold=is_ni,
               size=12 if is_ni else 11, color=NAVY if is_ni else INK,
               bg=CREAM if is_ni else None)
AVG_CONTRIB_REF = "'CVP Dashboard'!$C$8"
NI_REF = "'CVP Dashboard'!$C$7"

# Breakeven block
section_banner(ws4, 10, 2, 3, "BREAKEVEN")
be = [
    (11, "Breakeven # of offices (at avg contribution)",
         f"=IF({AVG_CONTRIB_REF}<=0,\"n/a\",ROUNDUP({FIXED_TOTAL_REF}/{AVG_CONTRIB_REF},0))", NUM0),
    (12, "Breakeven contribution required PER OFFICE (at current count)",
         f"=IF({OFFICE_COUNT_REF}=0,\"n/a\",{FIXED_TOTAL_REF}/{OFFICE_COUNT_REF})", USD),
    (13, "Margin of safety (offices above breakeven)",
         f"=IF({AVG_CONTRIB_REF}<=0,\"n/a\",{OFFICE_COUNT_REF}-ROUNDUP({FIXED_TOTAL_REF}/{AVG_CONTRIB_REF},0))", NUM0),
    (14, "Contribution margin coverage (contribution ÷ fixed)",
         f"=IF({FIXED_TOTAL_REF}=0,\"n/a\",{TOTAL_CONTRIB_REF}/{FIXED_TOTAL_REF})", PCT),
]
for r, label, formula, fmt in be:
    style_cell(ws4.cell(r, 2, label), align="left", border=box)
    style_cell(ws4.cell(r, 3, formula), number_format=fmt, align="right", border=box, bold=True, color=NAVY)

style_cell(ws4.cell(16, 2,
           "The per-office breakeven requirement (row 12) FALLS as offices are added — fixed costs are shared "
           "across more offices. The schedule below shows contribution stacking against the fixed-cost line."),
           align="left", wrap=True, color=GREY, italic=True)
ws4.merge_cells("B16:C17")

# Breakeven-shift schedule
sch_hdr = 19
section_banner(ws4, sch_hdr, 5, 8, "BREAKEVEN-SHIFT SCHEDULE  (as offices are added, at avg contribution)")
sh = sch_hdr + 1
sch_headers = ["# Offices", "Cumulative Contribution", "Fixed Costs", "Net Income"]
for i, h in enumerate(sch_headers):
    style_cell(ws4.cell(sh, 5 + i, h), bold=True, color=WHITE, bg=NAVY_DEEP, align="center", border=box, wrap=True, size=10)
ws4.row_dimensions[sh].height = 28

SCHED_N = 30
sch_first = sh + 1
for n in range(1, SCHED_N + 1):
    r = sch_first + n - 1
    style_cell(ws4.cell(r, 5, n), align="center", border=box, number_format=NUM0)
    style_cell(ws4.cell(r, 6, f"=E{r}*{AVG_CONTRIB_REF}"), number_format=USD, align="right", border=box)
    style_cell(ws4.cell(r, 7, f"={FIXED_TOTAL_REF}"), number_format=USD, align="right", border=box, color=GREY)
    style_cell(ws4.cell(r, 8, f"=F{r}-G{r}"), number_format=USD, align="right", border=box)
sch_last = sch_first + SCHED_N - 1

# Line chart: cumulative contribution vs fixed
chart = LineChart()
chart.title = "Contribution vs Fixed Costs by Office Count"
chart.style = 2
chart.height = 8
chart.width = 16
chart.y_axis.title = "Annual $"
chart.x_axis.title = "# Offices"
data = Reference(ws4, min_col=6, max_col=8, min_row=sh, max_row=sch_last)
cats = Reference(ws4, min_col=5, min_row=sch_first, max_row=sch_last)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
# color series: contribution navy, fixed gold, net income green
colors = [NAVY, GOLD_DEEP, OK_GREEN]
for s, col in zip(chart.series, colors):
    s.graphicalProperties.line.solidFill = col
    s.graphicalProperties.line.width = 28000
chart.x_axis.delete = False
chart.y_axis.delete = False
ws4.add_chart(chart, "B20")

# ============================================================================
# TAB 5 — New Office Evaluator
# ============================================================================
ws5 = wb.create_sheet("New Office Evaluator")
ws5.sheet_view.showGridLines = False
widths5 = {"A": 2.5, "B": 42, "C": 18, "D": 4, "E": 40, "F": 18}
for col, w in widths5.items():
    ws5.column_dimensions[col].width = w

style_cell(ws5.cell(1, 2, "New Office Evaluator"), bold=True, size=16, color=NAVY, align="left")
ws5.row_dimensions[1].height = 24
style_cell(ws5.cell(2, 2, "Enter a prospective office to price it off the tiered engine and see its marginal impact."),
           align="left", color=GREY, italic=True)

section_banner(ws5, 4, 2, 3, "PROSPECT INPUTS")
inputs5 = [
    (5, "Model (Access / Founders)", "Access", None),
    (6, "Base fee (blank = use model default)", "", USD),
    (7, "Number of accounts", 150, NUM0),
    (8, "Number of IARs (advisors)", 1, NUM0),
    (9, "Cost to serve (annual, variable)", 0, USD),
]
for r, label, val, fmt in inputs5:
    style_cell(ws5.cell(r, 2, label), align="left", border=box)
    c = ws5.cell(r, 3, val)
    style_cell(c, align="right" if fmt else "center", bg="DCE6F5", border=box,
               number_format=fmt if fmt else None)
dv_model5 = DataValidation(type="list", formula1='"Access,Founders"', allow_blank=True)
ws5.add_data_validation(dv_model5)
dv_model5.add(ws5.cell(5, 3))

# Computed outputs
section_banner(ws5, 11, 2, 3, "PRICING & CONTRIBUTION")
# modeled revenue for the prospect
base_expr = f'IF(C6="",IF(C5="Founders",{BF_FOUND_REF},{BF_ACCESS_REF}),C6)'
peracct_expr = f'IF(N(C7)=0,0,C7*LOOKUP(C7,{PA_LOOKUP_VEC},{PA_RESULT_VEC})*12)'
peradv_expr = f'MAX(N(C8)-1,0)*LOOKUP(N(C8),{AD_LOOKUP_VEC},{AD_RESULT_VEC})'
outputs5 = [
    (12, "Base platform fee", f"={base_expr}", USD),
    (13, "Per-account revenue (annual)", f"={peracct_expr}", USD),
    (14, "Per-advisor revenue (annual)", f"={peradv_expr}", USD),
    (15, "Total revenue to Rossby", f"={base_expr}+{peracct_expr}+{peradv_expr}", USD),
    (16, "Contribution (revenue − cost to serve)", f"={base_expr}+{peracct_expr}+{peradv_expr}-N(C9)", USD),
]
for r, label, formula, fmt in outputs5:
    style_cell(ws5.cell(r, 2, label), align="left", border=box)
    is_tot = r in (15, 16)
    style_cell(ws5.cell(r, 3, formula), number_format=fmt, align="right", border=box,
               bold=is_tot, color=NAVY if is_tot else INK, bg=CREAM if r == 16 else None)
PROSPECT_CONTRIB = "'New Office Evaluator'!$C$16"

section_banner(ws5, 18, 2, 3, "MARGINAL IMPACT ON FIRM")
impact5 = [
    (19, "Firm net income BEFORE this office", f"={NI_REF}", USD),
    (20, "Firm net income AFTER this office", f"={NI_REF}+{PROSPECT_CONTRIB}", USD),
    (21, "Change in net income", f"={PROSPECT_CONTRIB}", USD),
    (22, "Offices after adding this one", f"={OFFICE_COUNT_REF}+1", NUM0),
]
for r, label, formula, fmt in impact5:
    style_cell(ws5.cell(r, 2, label), align="left", border=box)
    style_cell(ws5.cell(r, 3, formula), number_format=fmt, align="right", border=box,
               bold=(r == 21), color=NAVY if r == 21 else INK)

style_cell(ws5.cell(24, 2,
           "Because Rossby is already near breakeven, most of a new office's contribution flows straight to "
           "net income. Add the office to the Register (Tab 3) to make it permanent."),
           align="left", wrap=True, color=GREY, italic=True)
ws5.merge_cells("B24:C25")

# ----------------------------------------------------------------------------
# Freeze header panes for the data tabs
# ----------------------------------------------------------------------------
ws2.freeze_panes = "A7"
ws3.freeze_panes = "B5"

# ----------------------------------------------------------------------------
OUT = "Rossby_CVP_Multi_Office_Model.xlsx"
import os
out_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), OUT)
wb.save(out_path)
print(f"Wrote {out_path}")
